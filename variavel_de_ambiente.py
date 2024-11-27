import pandas as pd

meu_dicionario = {
    "nome": ["João","jose", "marcos"],
    "idade": [21, 29, 32],
    "cidade": ["fortaleza", "são paulo", "minas gerais"]
    
}

tabela = pd.DataFrame.from_dict(meu_dicionario)
tabela.to_excel("tabela2.xlsx", index=False)

from openpyxl import Workbook
from openpyxl.styles import Alignment

arquivo = Workbook()
aba = arquivo.active
aba.title = "base de dados"

titulo = arquivo.active
titulo.title = "PLANILHA DE GASTOS"
titulo.merge_cells("A1:F1")
nomes_colunas = list(meu_dicionario.keys())

for num_coluna, nome_coluna in enumerate(meu_dicionario, start=1):
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    celula = aba.cell(row=2, column=num_coluna)
    celula.value = nome_coluna


arquivo.save("dados.xlsx")