from openpyxl import Workbook

# Cria uma nova pasta de trabalho e ativa a planilha
wb = Workbook()
ws = wb.active

# Renomeia a planilha
ws.title = "Exemplo Mesclar"

# Adiciona um valor antes de mesclar
ws["A1"] = "Texto mesclado"

# Mescla as células de A1 até C1
ws.merge_cells("A1:C1")

# Adiciona outro exemplo
ws["A2"] = "Outro exemplo"
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)

# Salva a planilha
wb.save("mesclar_celulas.xlsx")

print("Planilha criada com células mescladas!")
