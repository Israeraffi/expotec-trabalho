import pandas as pd

variavel = input('insira seus dados')

meu_dicionario = {
    
    "Janeiro": [variavel],
    "Fevereiro": [variavel],
    "Março": [variavel],
    "Abril": [variavel],
    "Maio": [variavel],
    "Junho": [variavel],
    "Julho": [variavel],
    "Agosto": [variavel],
    "Setembro": [variavel],
    "Outubro": [variavel],
    "Novembro": [variavel],
    "Dezembro": [variavel]
    
}



tabela = pd.DataFrame.from_dict(meu_dicionario)
tabela.to_excel("tabela2.xlsx", index=False)

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

arquivo = Workbook()
aba = arquivo.active
aba.title = "base de dados"

mes = arquivo.active
mes["A2"] = "Mês"
mes["A2"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
mes["A2"].font = Font(bold=16)

titulo = arquivo.active
titulo["A1"] = "PLANILHA DE GASTOS"
titulo.title = "PLANILHA DE GASTOS"
titulo.merge_cells("A1:M1")
nomes_colunas = list(meu_dicionario.keys())

for num_coluna, nome_coluna in enumerate(meu_dicionario, start=1):
    titulo["A1"].alignment = Alignment(horizontal="center", vertical="center")
    titulo.row_dimensions[1].height = 30
    titulo["A1"].font = Font(color="00FF00")
    titulo["A1"].font = Font(bold=20)
    titulo["A1"].fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    celula = aba.cell(row=2, column=num_coluna+1)
    celula.value = nome_coluna


#AUMENTAR O TAMANHO DAS CELULAS
aba.column_dimensions["B"].width = 12
aba.column_dimensions["C"].width = 12
aba.column_dimensions["D"].width = 12
aba.column_dimensions["E"].width = 12
aba.column_dimensions["F"].width = 12
aba.column_dimensions["G"].width = 12
aba.column_dimensions["H"].width = 12
aba.column_dimensions["I"].width = 12
aba.column_dimensions["J"].width = 12
aba.column_dimensions["K"].width = 12
aba.column_dimensions["L"].width = 12
aba.column_dimensions["M"].width = 12

#CENTRALIZAR
aba["B2"].alignment = Alignment(horizontal="center", vertical="center")
aba["C2"].alignment = Alignment(horizontal="center", vertical="center")
aba["D2"].alignment = Alignment(horizontal="center", vertical="center")
aba["E2"].alignment = Alignment(horizontal="center", vertical="center")
aba["F2"].alignment = Alignment(horizontal="center", vertical="center")
aba["G2"].alignment = Alignment(horizontal="center", vertical="center")
aba["H2"].alignment = Alignment(horizontal="center", vertical="center")
aba["I2"].alignment = Alignment(horizontal="center", vertical="center")
aba["J2"].alignment = Alignment(horizontal="center", vertical="center")
aba["K2"].alignment = Alignment(horizontal="center", vertical="center")
aba["L2"].alignment = Alignment(horizontal="center", vertical="center")
aba["M2"].alignment = Alignment(horizontal="center", vertical="center")


#MUDAR COR DA CELULA
aba["B2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["C2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["D2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["E2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["F2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["G2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["H2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["I2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["J2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["K2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["L2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["M2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")


#AUMENTAR O TAMANHO DA FONTE
aba["B2"].font = Font(bold=16)
aba["C2"].font = Font(bold=16)
aba["D2"].font = Font(bold=16)
aba["E2"].font = Font(bold=16)
aba["F2"].font = Font(bold=16)
aba["G2"].font = Font(bold=16)
aba["H2"].font = Font(bold=16)
aba["I2"].font = Font(bold=16)
aba["J2"].font = Font(bold=16)
aba["K2"].font = Font(bold=16)
aba["L2"].font = Font(bold=16)
aba["M2"].font = Font(bold=16)

arquivo.save("dados.xlsx")
