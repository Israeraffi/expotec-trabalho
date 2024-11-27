from openpyxl import Workbook
from openpyxl.styles import Alignment

# Cria uma nova pasta de trabalho e ativa a planilha
wb = Workbook()
ws = wb.active

# Renomeia a planilha
ws.title = "Centralizar Texto"

# Adiciona dados na célula
ws["A1"] = "Texto centralizado"

# Define o alinhamento central (horizontal e vertical)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Ajusta a largura da coluna e altura da linha (opcional)
ws.column_dimensions["A"].width = 20
ws.row_dimensions[1].height = 30

# Salva o arquivo
wb.save("texto_centralizado.xlsx")

print("Planilha criada com texto centralizado!")
