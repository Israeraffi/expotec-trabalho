import pandas as pd

variavel = input('insira seus dados')

meu_dicionario = {
    "Janeiro": [variavel],
    "Fevereiro": [variavel],
    "Março": [variavel],
    "Abril": [variavel],
    "Maio": [variavel],
    "Junho": [variavel],
    "Julho": [variavel],
    "Agosto": [variavel],
    "Setembro": [variavel],
    "Outubro": [variavel],
    "Novembro": [variavel],
    "Dezembro": [variavel]
    
}

tabela = pd.DataFrame.from_dict(meu_dicionario)
tabela.to_excel("tabela2.xlsx", index=False)

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

arquivo = Workbook()
aba = arquivo.active
aba.title = "base de dados"

titulo = arquivo.active
titulo["A1"] = "PLANILHA DE GASTOS"
titulo.title = "PLANILHA DE GASTOS"
titulo.merge_cells("A1:M1")
nomes_colunas = list(meu_dicionario.keys())

for num_coluna, nome_coluna in enumerate(meu_dicionario, start=1):
    titulo["A1"].alignment = Alignment(horizontal="center", vertical="center")
    titulo.row_dimensions[1].height = 30
    titulo["A1"].font = Font(color="00FF00")
    titulo["A1"].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    celula = aba.cell(row=2, column=num_coluna+1)
    celula.value = nome_coluna


arquivo.save("dados.xlsx")
