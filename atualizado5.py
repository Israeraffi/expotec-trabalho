import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

arquivo = Workbook()
aba = arquivo.active
aba.title = "base de dados"


meu_dicionario = {
    
    "Janeiro": [""],
    "Fevereiro": [""],
    "Março": [""],
    "Abril": [""],
    "Maio": [""],
    "Junho": [""],
    "Julho": [""],
    "Agosto": [""],
    "Setembro": [""],
    "Outubro": [""],
    "Novembro": [""],
    "Dezembro": [""]
    
}

dicionario = {
    
    "Janeiro": [""],
    "Fevereiro": [""],
    "Março": [""],
    "Abril": [""],
    "Maio": [""],
    "Junho": [""],
    "Julho": [""],
    "Agosto": [""],
    "Setembro": [""],
    "Outubro": [""],
    "Novembro": [""],
    "Dezembro": [""]
    
}


tabela = pd.DataFrame.from_dict(meu_dicionario)
tabela.to_excel("tabela2.xlsx", index=False)

despesa = arquivo.active
despesa["A11"] = " Despesas"
despesa["A11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
despesa["A11"].font = Font(bold=16)

mes = arquivo.active
mes["A2"] = "Mês"
mes["A2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
mes["A2"].font = Font(bold=16)

mercado = arquivo.active
mercado["A12"] = "Mercado"
mercado ["A12"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
mercado ["A12"].font = Font(bold=16)

aluguel = arquivo.active
aluguel["A13"] = "Aluguel"
aluguel ["A13"].font = Font(bold=16)

transporte = arquivo.active
transporte["A14"] = "Transporte"
transporte ["A14"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
transporte ["A14"].font = Font(bold=16)

agua = arquivo.active
agua ["A15"] = "Água"
agua ["A15"].font = Font(bold=16)

luz = arquivo.active
luz["A16"] = "Luz"
luz ["A16"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
luz ["A16"].font = Font(bold=16)

internet = arquivo.active
internet["A17"] = "internet"
internet ["A17"].font = Font(bold=16)

total = arquivo.active
total["A18"] = "TOTAL"
total["A18"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
total["A18"].font = Font(bold=16)

gestao = arquivo.active
gestao ["A6"] = "Salário"
gestao ["A6"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
gestao ["A6"].font = Font(bold=16)

investimento = arquivo.active
investimento ["A7"] = "Investimentos"
investimento ["A7"].font = Font(bold=16)

renda_extra = arquivo.active
renda_extra["A8"] = "Renda extra"
renda_extra["A8"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
renda_extra["A8"].font = Font(bold=16)

totall = arquivo.active
totall["B3"] = "=B9-B18"

totall2 = arquivo.active
totall2["C3"] = "=C9-BC18"

totall3 = arquivo.active
totall3["D3"] = "=D9-D18"

totall4 = arquivo.active
totall4["E3"] = "=E9-E18"

totall5 = arquivo.active
totall5["F3"] = "=F9-F18"

totall6 = arquivo.active
totall6["G3"] = "=G9-G18"

totall7 = arquivo.active
totall7["H3"] = "=H9-H18"

totall8 = arquivo.active
totall8["I3"] = "=I9-I18"

totall9 = arquivo.active
totall9["J3"] = "=J9-J18"

totall10 = arquivo.active
totall10["K3"] = "=K9-K18"

totall11 = arquivo.active
totall11["L3"] = "=L9-L18"

totall12 = arquivo.active
totall12["M3"] = "=M9-M18"
totall12.number_format = 'R$ #,##0.00'

receitas = arquivo.active
receitas["A5"] = "Receitas"
receitas["A5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
receitas["A5"].font = Font(bold=16)

total_receitas = arquivo.active
total_receitas["A9"] = "TOTAL"
total_receitas["A9"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
total_receitas["A9"].font = Font(bold=16)

sub_mes = arquivo.active
sub_mes["A3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")

titulo = arquivo.active
titulo["A1"] = "PLANILHA DE GASTOS"
titulo.title = "PLANILHA DE GASTOS"
titulo.merge_cells("A1:M1")
nomes_colunas = list(meu_dicionario.keys())

nomes_das_colunas = list(dicionario.keys())

for num_coluna, nome_coluna in enumerate(meu_dicionario, start=1):
    titulo["A1"].alignment = Alignment(horizontal="center", vertical="center")
    titulo.row_dimensions[1].height = 30
    titulo["A1"].font = Font(color="00FF00")
    titulo["A1"].font = Font(bold=20)
    titulo["A1"].fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    celula = aba.cell(row=2, column=num_coluna+1)
    celula.value = nome_coluna

for num_coluna, nome_coluna in enumerate(dicionario, start=1):
        celula = aba.cell(row=5, column=num_coluna+1)
        celula.value = nome_coluna

for num_coluna, nome_coluna in enumerate(dicionario, start=1):
        celula = aba.cell(row=11, column=num_coluna+1)
        celula.value = nome_coluna

#AUMENTAR O TAMANHO DAS CELULAS
aba.column_dimensions["A"].width = 14
aba.column_dimensions["B"].width = 14
aba.column_dimensions["C"].width = 14
aba.column_dimensions["D"].width = 14
aba.column_dimensions["E"].width = 14
aba.column_dimensions["F"].width = 14
aba.column_dimensions["G"].width = 14
aba.column_dimensions["H"].width = 14
aba.column_dimensions["I"].width = 14
aba.column_dimensions["J"].width = 14
aba.column_dimensions["K"].width = 14
aba.column_dimensions["L"].width = 14
aba.column_dimensions["M"].width = 14
aba.column_dimensions["N"].width = 14


#CENTRALIZAR LINHA 5
aba["B5"].alignment = Alignment(horizontal="center", vertical="center")
aba["C5"].alignment = Alignment(horizontal="center", vertical="center")
aba["D5"].alignment = Alignment(horizontal="center", vertical="center")
aba["E5"].alignment = Alignment(horizontal="center", vertical="center")
aba["F5"].alignment = Alignment(horizontal="center", vertical="center")
aba["G5"].alignment = Alignment(horizontal="center", vertical="center")
aba["H5"].alignment = Alignment(horizontal="center", vertical="center")
aba["I5"].alignment = Alignment(horizontal="center", vertical="center")
aba["J5"].alignment = Alignment(horizontal="center", vertical="center")
aba["K5"].alignment = Alignment(horizontal="center", vertical="center")
aba["L5"].alignment = Alignment(horizontal="center", vertical="center")
aba["M5"].alignment = Alignment(horizontal="center", vertical="center")


#MUDAR COR DA CELULA LINHA 5
aba["B5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["C5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["D5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["E5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["F5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["G5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["H5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["I5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["J5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["K5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["L5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["M5"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")


#AUMENTAR O TAMANHO DA FONTE LINHA 5
aba["B5"].font = Font(bold=16)
aba["C5"].font = Font(bold=16)
aba["D5"].font = Font(bold=16)
aba["E5"].font = Font(bold=16)
aba["F5"].font = Font(bold=16)
aba["G5"].font = Font(bold=16)
aba["H5"].font = Font(bold=16)
aba["I5"].font = Font(bold=16)
aba["J5"].font = Font(bold=16)
aba["K5"].font = Font(bold=16)
aba["L5"].font = Font(bold=16)
aba["M5"].font = Font(bold=16)


#CENTRALIZAR LINHA 2
aba["B2"].alignment = Alignment(horizontal="center", vertical="center")
aba["C2"].alignment = Alignment(horizontal="center", vertical="center")
aba["D2"].alignment = Alignment(horizontal="center", vertical="center")
aba["E2"].alignment = Alignment(horizontal="center", vertical="center")
aba["F2"].alignment = Alignment(horizontal="center", vertical="center")
aba["G2"].alignment = Alignment(horizontal="center", vertical="center")
aba["H2"].alignment = Alignment(horizontal="center", vertical="center")
aba["I2"].alignment = Alignment(horizontal="center", vertical="center")
aba["J2"].alignment = Alignment(horizontal="center", vertical="center")
aba["K2"].alignment = Alignment(horizontal="center", vertical="center")
aba["L2"].alignment = Alignment(horizontal="center", vertical="center")
aba["M2"].alignment = Alignment(horizontal="center", vertical="center")


#MUDAR COR DA CELULA LINHA 2
aba["B2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["C2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["D2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["E2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["F2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["G2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["H2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["I2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["J2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["K2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["L2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["M2"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")


#AUMENTAR O TAMANHO DA FONTE LINHA 2
aba["B2"].font = Font(bold=16)
aba["C2"].font = Font(bold=16)
aba["D2"].font = Font(bold=16)
aba["E2"].font = Font(bold=16)
aba["F2"].font = Font(bold=16)
aba["G2"].font = Font(bold=16)
aba["H2"].font = Font(bold=16)
aba["I2"].font = Font(bold=16)
aba["J2"].font = Font(bold=16)
aba["K2"].font = Font(bold=16)
aba["L2"].font = Font(bold=16)
aba["M2"].font = Font(bold=16)

#COR LINHA 9
aba["B11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["C11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["D11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["E11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["F11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["G11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["H11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["I11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["J11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["K11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["L11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["M11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")


aba["B11"].alignment = Alignment(horizontal="center", vertical="center")
aba["C11"].alignment = Alignment(horizontal="center", vertical="center")
aba["D11"].alignment = Alignment(horizontal="center", vertical="center")
aba["E11"].alignment = Alignment(horizontal="center", vertical="center")
aba["F11"].alignment = Alignment(horizontal="center", vertical="center")
aba["G11"].alignment = Alignment(horizontal="center", vertical="center")
aba["H11"].alignment = Alignment(horizontal="center", vertical="center")
aba["I11"].alignment = Alignment(horizontal="center", vertical="center")
aba["J11"].alignment = Alignment(horizontal="center", vertical="center")
aba["K11"].alignment = Alignment(horizontal="center", vertical="center")
aba["L11"].alignment = Alignment(horizontal="center", vertical="center")
aba["M11"].alignment = Alignment(horizontal="center", vertical="center")


#MUDAR COR DA CELULA LINHA 5
aba["B11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["C11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["D11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["E11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["F11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["G11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["H11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["I11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["J11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["K11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["L11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["M11"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")


#AUMENTAR O TAMANHO DA FONTE LINHA 5
aba["B11"].font = Font(bold=16)
aba["C11"].font = Font(bold=16)
aba["D11"].font = Font(bold=16)
aba["E11"].font = Font(bold=16)
aba["F11"].font = Font(bold=16)
aba["G11"].font = Font(bold=16)
aba["H11"].font = Font(bold=16)
aba["I11"].font = Font(bold=16)
aba["J11"].font = Font(bold=16)
aba["K11"].font = Font(bold=16)
aba["L11"].font = Font(bold=16)
aba["M11"].font = Font(bold=16)

#ATUALIZADO
aba["B3"].alignment = Alignment(horizontal="center", vertical="center")
aba["C3"].alignment = Alignment(horizontal="center", vertical="center")
aba["D3"].alignment = Alignment(horizontal="center", vertical="center")
aba["E3"].alignment = Alignment(horizontal="center", vertical="center")
aba["F3"].alignment = Alignment(horizontal="center", vertical="center")
aba["G3"].alignment = Alignment(horizontal="center", vertical="center")
aba["H3"].alignment = Alignment(horizontal="center", vertical="center")
aba["I3"].alignment = Alignment(horizontal="center", vertical="center")
aba["J3"].alignment = Alignment(horizontal="center", vertical="center")
aba["K3"].alignment = Alignment(horizontal="center", vertical="center")
aba["L3"].alignment = Alignment(horizontal="center", vertical="center")
aba["M3"].alignment = Alignment(horizontal="center", vertical="center")



aba["B3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["C3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["D3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["E3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["F3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["G3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["H3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["I3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["J3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["K3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["L3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
aba["M3"].fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")


#AUMENTAR O TAMANHO DA FONTE LINHA 3
aba["B3"].font = Font(bold=16)
aba["C3"].font = Font(bold=16)
aba["D3"].font = Font(bold=16)
aba["E3"].font = Font(bold=16)
aba["F3"].font = Font(bold=16)
aba["G3"].font = Font(bold=16)
aba["H3"].font = Font(bold=16)
aba["I3"].font = Font(bold=16)
aba["J3"].font = Font(bold=16)
aba["K3"].font = Font(bold=16)
aba["L3"].font = Font(bold=16)
aba["M3"].font = Font(bold=16).alignment = Alignment(horizontal="center", vertical="center")
for col in range(2, 14):  
    celula_total = aba.cell(row=9, column=col)
    celula_total.value = f"=SUM({chr(64 + col)}6:{chr(64 + col)}8)"
    celula_total.alignment = Alignment(horizontal="center", vertical="center")
    celula_total.font = Font(bold=16)
    celula_total.fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
    celula_total.number_format = 'R$ #,##0.00'

for col1 in range(2, 14):  
    celula_total1 = aba.cell(row=18, column=col1)
    celula_total1.value = f"=SUM({chr(64 + col1)}12:{chr(64 + col1)}17)"
    celula_total1.font = Font(bold=16)
    celula_total1.fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
    celula_total1.number_format = 'R$ #,##0.00'
    celula_total1.alignment = Alignment(horizontal="center", vertical="center")




for col in range(1, 18 + 1): 
    celula = aba.cell(row=1, column=col)
    celula.alignment = Alignment(horizontal="center", vertical="center")


arquivo.save("dados.xlsx")
